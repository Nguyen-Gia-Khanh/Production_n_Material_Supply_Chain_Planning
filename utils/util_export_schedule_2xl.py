"""Utilities for exporting solved production schedules to Excel."""

from pathlib import Path
from typing import Any, Mapping, Sequence

import pandas as pd
import pulp as pl
from openpyxl import load_workbook


def export_daily_production_schedule(
    *,
    model: pl.LpProblem,
    production: Mapping[tuple[Any, Any], pl.LpVariable],
    products: Sequence[Any],
    days: Sequence[Any],
    day_to_date: Mapping[Any, Any],
    planning_input: str | Path,
    sheet_name: str = "Production_Schedule",
) -> pd.DataFrame:
    """Write solved daily production quantities into an existing XLSM template.

    The template must store dates on row 3 from column B onward and products
    in column A from row 4 onward. Existing formatting and VBA are preserved.
    """
    status_name = pl.LpStatus[model.status]
    if status_name not in {"Optimal", "Feasible"}:
        raise RuntimeError(
            f"Cannot export schedule. Solver status: {status_name}"
        )

    products = list(products)
    days = list(days)

    production_result = pd.DataFrame(
        {
            day_to_date[day]: [
                int(round(pl.value(production[product, day])))
                for product in products
            ]
            for day in days
        },
        index=products,
    )

    workbook = load_workbook(planning_input, keep_vba=True)
    try:
        worksheet = workbook[sheet_name]

        template_products = [
            worksheet.cell(row=4 + position, column=1).value
            for position in range(len(products))
        ]

        template_dates = [
            pd.Timestamp(
                worksheet.cell(row=3, column=2 + position).value
            ).normalize()
            for position in range(len(days))
        ]

        model_dates = [
            pd.Timestamp(day_to_date[day]).normalize()
            for day in days
        ]

        if template_products != products:
            raise ValueError(
                f"Product order in {sheet_name} does not match "
                "the optimization model."
            )

        if template_dates != model_dates:
            raise ValueError(
                f"Dates in {sheet_name} do not match "
                "the optimization horizon."
            )

        for row_number, product in enumerate(products, start=4):
            for column_number, day in enumerate(days, start=2):
                worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).value = int(
                    production_result.loc[product, day_to_date[day]]
                )

        workbook.save(planning_input)
    finally:
        workbook.close()

    print(
        f"Schedule exported to "
        f"{planning_input} → {sheet_name}!B4"
    )

    return production_result
